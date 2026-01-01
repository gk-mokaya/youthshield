from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum
from django.db.models.functions import TruncMonth
from datetime import datetime, timedelta
from django.utils import timezone
import json

from core.models import WebsiteSetting, CoreValue, BoardMember, ExecutiveCommittee, ContactMessage
from users.models import CustomUser
from testimonials.models import Testimonial
from programs.models import Program, Service, Objective
from donations.models import Donation

def is_staff_member(user):
    return user.is_staff_member()

@login_required
@user_passes_test(is_staff_member)
def dashboard(request):
    """Main staff dashboard view"""
    if not request.user or not request.user.is_authenticated:
        from django.http import HttpResponseRedirect
        from django.urls import reverse
        return HttpResponseRedirect(reverse('login'))

    context = {
        'user': request.user,
        'total_users': CustomUser.objects.filter(user_type='user').count(),
        'total_staff': CustomUser.objects.filter(user_type='staff').count(),
        'total_donations': Donation.objects.count(),
        'total_testimonials': Testimonial.objects.filter(status='approved').count(),
        'total_programs': Program.objects.filter(is_active=True).count(),
        'unresolved_messages': ContactMessage.objects.filter(resolved=False).count(),
        'recent_messages': ContactMessage.objects.filter(is_seen=False).order_by('-created_at')[:5],
        'recent_donations': Donation.objects.order_by('-created_at')[:5],
    }
    return render(request, 'staff_dashboard/dashboard.html', context)

@login_required
@user_passes_test(is_staff_member)
def analytics_data(request):
    """API endpoint for analytics chart data"""
    print(f"Analytics data requested by user: {request.user}, is_staff: {request.user.is_staff_member()}")
    # User registrations over last 12 months
    end_date = timezone.now()
    start_date = end_date - timedelta(days=365)

    user_registrations = (
        CustomUser.objects
        .filter(created_at__gte=start_date, user_type='user')
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )

    # Donations over last 12 months
    donation_data = (
        Donation.objects
        .filter(created_at__gte=start_date)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'), total_amount=Sum('amount'))
        .order_by('month')
    )

    # Testimonials status distribution
    testimonial_status = (
        Testimonial.objects
        .values('status')
        .annotate(count=Count('id'))
    )

    # Contact messages status
    message_status = (
        ContactMessage.objects
        .values('resolved')
        .annotate(count=Count('id'))
    )

    # Programs data
    programs_data = list(Program.objects.values('title', 'is_active'))

    # Additional analytics data
    # Donation payment methods distribution
    donation_payment_methods = (
        Donation.objects
        .values('payment_method')
        .annotate(count=Count('id'), total_amount=Sum('amount'))
        .order_by('payment_method')
    )

    # Donation status distribution
    donation_status = (
        Donation.objects
        .values('status')
        .annotate(count=Count('id'), total_amount=Sum('amount'))
        .order_by('status')
    )

    # Testimonial ratings distribution
    testimonial_ratings = (
        Testimonial.objects
        .values('rating')
        .annotate(count=Count('id'))
        .order_by('rating')
    )

    # Program categories distribution (if category field exists)
    try:
        program_categories = list(
            Program.objects
            .values('category')
            .annotate(count=Count('id'), active_count=Count('id', filter=Q(is_active=True)))
            .order_by('category')
        )
    except:
        # Fallback if category field doesn't exist
        program_categories = []

    # Monthly revenue trend (total donation amounts)
    monthly_revenue = (
        Donation.objects
        .filter(status='completed')
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(total_revenue=Sum('amount'))
        .order_by('month')
    )

    # User activity - last 30 days active users
    thirty_days_ago = end_date - timedelta(days=30)
    active_users = CustomUser.objects.filter(last_login__gte=thirty_days_ago).count()

    # Recent activity summary
    recent_activity = {
        'new_users_last_7_days': CustomUser.objects.filter(created_at__gte=end_date - timedelta(days=7)).count(),
        'new_donations_last_7_days': Donation.objects.filter(created_at__gte=end_date - timedelta(days=7)).count(),
        'new_testimonials_last_7_days': Testimonial.objects.filter(created_at__gte=end_date - timedelta(days=7)).count(),
        'messages_resolved_last_7_days': ContactMessage.objects.filter(
            resolved=True,
            created_at__gte=end_date - timedelta(days=7)
        ).count(),
    }

    return JsonResponse({
        'user_registrations': list(user_registrations),
        'donation_data': list(donation_data),
        'testimonial_status': list(testimonial_status),
        'message_status': list(message_status),
        'programs_data': programs_data,
        'donation_payment_methods': list(donation_payment_methods),
        'donation_status': list(donation_status),
        'testimonial_ratings': list(testimonial_ratings),
        'program_categories': program_categories,
        'monthly_revenue': list(monthly_revenue),
        'active_users': active_users,
        'recent_activity': recent_activity,
    })

# User Management Views
@login_required
@user_passes_test(is_staff_member)
def user_management(request):
    """User management page"""
    users = CustomUser.objects.all().order_by('-created_at')
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Calculate analytics data
    total_users = users.count()
    active_users = users.filter(is_active=True).count()
    staff_users = users.filter(user_type='staff').count()
    new_users_today = users.filter(created_at__date=timezone.now().date()).count()

    context = {
        'page_obj': page_obj,
        'total_users': total_users,
        'active_users': active_users,
        'staff_users': staff_users,
        'new_users_today': new_users_today,
    }
    return render(request, 'staff_dashboard/user_management.html', context)

@login_required
@user_passes_test(is_staff_member)
@require_POST
def create_user(request):
    """Create new user via AJAX"""
    try:
        data = json.loads(request.body)
        user = CustomUser.objects.create_user(
            username=data['username'],
            email=data['email'],
            password=data['password'],
            first_name=data.get('first_name', ''),
            last_name=data.get('last_name', ''),
            user_type=data.get('user_type', 'user'),
            phone_number=data.get('phone_number', ''),
            address=data.get('address', ''),
        )
        return JsonResponse({
            'success': True,
            'message': 'User created successfully',
            'user_id': user.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def update_user(request, user_id):
    """Update user via AJAX"""
    try:
        user = get_object_or_404(CustomUser, id=user_id)
        data = json.loads(request.body)

        user.username = data.get('username', user.username)
        user.email = data.get('email', user.email)
        user.first_name = data.get('first_name', user.first_name)
        user.last_name = data.get('last_name', user.last_name)
        user.user_type = data.get('user_type', user.user_type)
        user.phone_number = data.get('phone_number', user.phone_number)
        user.address = data.get('address', user.address)
        user.is_active = data.get('is_active', user.is_active)

        if data.get('password'):
            user.set_password(data['password'])

        user.save()

        return JsonResponse({
            'success': True,
            'message': 'User updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_user(request, user_id):
    """Delete user via AJAX"""
    try:
        user = get_object_or_404(CustomUser, id=user_id)
        user.delete()
        return JsonResponse({
            'success': True,
            'message': 'User deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Website Settings Management
@login_required
@user_passes_test(is_staff_member)
def website_settings(request):
    """Website settings management"""
    settings = WebsiteSetting.objects.first()
    if not settings:
        settings = WebsiteSetting.objects.create(
            name="Youth Shield Foundation",
            mission="Default mission",
            vision="Default vision",
            contact_email="info@youthshield.org",
            contact_phone="+1234567890",
            address="Default address"
        )

    if request.method == 'POST':
        # Handle form submission
        settings.name = request.POST.get('name', settings.name)
        settings.mission = request.POST.get('mission', settings.mission)
        settings.vision = request.POST.get('vision', settings.vision)
        settings.contact_email = request.POST.get('contact_email', settings.contact_email)
        settings.contact_phone = request.POST.get('contact_phone', settings.contact_phone)
        settings.address = request.POST.get('address', settings.address)
        settings.tagline = request.POST.get('tagline', settings.tagline)
        settings.description = request.POST.get('description', settings.description)

        # Handle social media links
        settings.facebook_url = request.POST.get('facebook_url')
        settings.twitter_url = request.POST.get('twitter_url')
        settings.instagram_url = request.POST.get('instagram_url')
        settings.linkedin_url = request.POST.get('linkedin_url')
        settings.tiktok_url = request.POST.get('tiktok_url')
        settings.whatsapp_channel_url = request.POST.get('whatsapp_channel_url')

        if 'logo' in request.FILES:
            settings.logo = request.FILES['logo']

        settings.save()
        messages.success(request, 'Website settings updated successfully!')
        return redirect('staff_dashboard:website_settings')

    context = {'settings': settings}
    return render(request, 'staff_dashboard/website_settings.html', context)

# Core Values Management
@login_required
@user_passes_test(is_staff_member)
def core_values_management(request):
    """Core values management"""
    core_values = CoreValue.objects.all().order_by('order')
    active_values_count = core_values.filter(is_active=True).count()
    context = {'core_values': core_values, 'active_values_count': active_values_count}
    return render(request, 'staff_dashboard/core_values.html', context)

@login_required
@user_passes_test(is_staff_member)
@require_POST
def create_core_value(request):
    """Create core value via AJAX"""
    try:
        data = json.loads(request.body)
        value = CoreValue.objects.create(
            name=data['name'],
            description=data['description'],
            icon_class=data.get('icon_class', 'fas fa-star'),
            order=data.get('order', 0),
            is_active=data.get('is_active', True)
        )
        return JsonResponse({
            'success': True,
            'message': 'Core value created successfully',
            'value_id': value.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def update_core_value(request, value_id):
    """Update core value via AJAX"""
    try:
        value = get_object_or_404(CoreValue, id=value_id)
        data = json.loads(request.body)

        value.name = data.get('name', value.name)
        value.description = data.get('description', value.description)
        value.icon_class = data.get('icon_class', value.icon_class)
        value.order = data.get('order', value.order)
        value.is_active = data.get('is_active', value.is_active)
        value.save()

        return JsonResponse({
            'success': True,
            'message': 'Core value updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def toggle_core_value(request, value_id):
    """Toggle core value active status via AJAX"""
    try:
        value = get_object_or_404(CoreValue, id=value_id)
        data = json.loads(request.body)
        value.is_active = data.get('is_active', not value.is_active)
        value.save()
        return JsonResponse({
            'success': True,
            'message': f'Core value {"activated" if value.is_active else "deactivated"} successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_core_value(request, value_id):
    """Delete core value via AJAX"""
    try:
        value = get_object_or_404(CoreValue, id=value_id)
        value.delete()
        return JsonResponse({
            'success': True,
            'message': 'Core value deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Board Members Management
@login_required
@user_passes_test(is_staff_member)
def board_members_management(request):
    """Board members management"""
    members = BoardMember.objects.all().order_by('order')
    active_members_count = members.filter(is_active=True).count()
    context = {'members': members, 'active_members_count': active_members_count}
    return render(request, 'staff_dashboard/board_members.html', context)

@login_required
@user_passes_test(is_staff_member)
@require_POST
def create_board_member(request):
    """Create board member via AJAX"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            # Handle FormData
            data = request.POST.dict()
            data['is_active'] = data.get('is_active') == 'on'
            data['order'] = int(data.get('order', 0))

        member = BoardMember.objects.create(
            name=data['name'],
            position=data['position'],
            bio=data.get('bio', ''),
            order=data.get('order', 0),
            is_active=data.get('is_active', True)
        )

        # Handle photo upload
        if 'photo' in request.FILES:
            member.photo = request.FILES['photo']
            member.save()

        return JsonResponse({
            'success': True,
            'message': 'Board member created successfully',
            'member_id': member.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def update_board_member(request, member_id):
    """Update board member via AJAX"""
    try:
        member = get_object_or_404(BoardMember, id=member_id)

        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            # Handle FormData
            data = request.POST.dict()
            data['is_active'] = data.get('is_active') == 'on'
            data['order'] = int(data.get('order', member.order))

        member.name = data.get('name', member.name)
        member.position = data.get('position', member.position)
        member.bio = data.get('bio', member.bio)
        member.order = data.get('order', member.order)
        member.is_active = data.get('is_active', member.is_active)

        # Handle photo upload
        if 'photo' in request.FILES:
            member.photo = request.FILES['photo']

        member.save()

        return JsonResponse({
            'success': True,
            'message': 'Board member updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def toggle_board_member(request, member_id):
    """Toggle board member active status via AJAX"""
    try:
        member = get_object_or_404(BoardMember, id=member_id)
        data = json.loads(request.body)
        member.is_active = data.get('is_active', not member.is_active)
        member.save()
        return JsonResponse({
            'success': True,
            'message': f'Board member {"activated" if member.is_active else "deactivated"} successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_board_member(request, member_id):
    """Delete board member via AJAX"""
    try:
        member = get_object_or_404(BoardMember, id=member_id)
        member.delete()
        return JsonResponse({
            'success': True,
            'message': 'Board member deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Executive Committee Management
@login_required
@user_passes_test(is_staff_member)
def executive_committee_management(request):
    """Executive committee management"""
    members = ExecutiveCommittee.objects.all().order_by('display_order')
    active_members = members.filter(is_active=True)
    positions_count = members.values('position').distinct().count()
    active_members_count = active_members.count()

    context = {
        'members': members,
        'active_members': active_members,
        'positions_count': positions_count,
        'active_members_count': active_members_count,
    }
    return render(request, 'staff_dashboard/executive_committee.html', context)

@login_required
@user_passes_test(is_staff_member)
@require_POST
def create_executive_member(request):
    """Create executive member via AJAX"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            # Handle FormData
            data = request.POST.dict()
            data['is_active'] = data.get('is_active') == 'on'
            data['display_order'] = int(data.get('display_order', 0))

        member = ExecutiveCommittee.objects.create(
            name=data['name'],
            position=data['position'],
            bio=data.get('bio'),
            email=data.get('email'),
            phone=data.get('phone'),
            display_order=data.get('display_order', 0),
            is_active=data.get('is_active', True)
        )

        # Handle photo upload
        if 'photo' in request.FILES:
            member.photo = request.FILES['photo']
            member.save()

        return JsonResponse({
            'success': True,
            'message': 'Executive member created successfully',
            'member_id': member.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def update_executive_member(request, member_id):
    """Update executive member via AJAX"""
    try:
        member = get_object_or_404(ExecutiveCommittee, id=member_id)

        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            # Handle FormData
            data = request.POST.dict()
            data['is_active'] = data.get('is_active') == 'on'
            data['display_order'] = int(data.get('display_order', member.display_order))

        member.name = data.get('name', member.name)
        member.position = data.get('position', member.position)
        member.bio = data.get('bio', member.bio)
        member.email = data.get('email', member.email)
        member.phone = data.get('phone', member.phone)
        member.display_order = data.get('display_order', member.display_order)
        member.is_active = data.get('is_active', member.is_active)

        # Handle photo upload
        if 'photo' in request.FILES:
            member.photo = request.FILES['photo']

        member.save()

        return JsonResponse({
            'success': True,
            'message': 'Executive member updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def toggle_executive_member(request, member_id):
    """Toggle executive member active status via AJAX"""
    try:
        member = get_object_or_404(ExecutiveCommittee, id=member_id)
        data = json.loads(request.body)
        member.is_active = data.get('is_active', not member.is_active)
        member.save()
        return JsonResponse({
            'success': True,
            'message': f'Executive member {"activated" if member.is_active else "deactivated"} successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_executive_member(request, member_id):
    """Delete executive member via AJAX"""
    try:
        member = get_object_or_404(ExecutiveCommittee, id=member_id)
        member.delete()
        return JsonResponse({
            'success': True,
            'message': 'Executive member deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Contact Messages Management
@login_required
@user_passes_test(is_staff_member)
def contact_messages_management(request):
    """Contact messages management"""
    messages_list = ContactMessage.objects.all().order_by('-created_at')
    paginator = Paginator(messages_list, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_messages': messages_list.count(),
        'unresolved_count': messages_list.filter(resolved=False).count(),
        'unread_count': messages_list.filter(is_seen=False).count(),
    }
    return render(request, 'staff_dashboard/contact_messages.html', context)

@login_required
@user_passes_test(is_staff_member)
def view_message(request, message_id):
    """View message details via AJAX"""
    try:
        message = get_object_or_404(ContactMessage, id=message_id)
        return JsonResponse({
            'success': True,
            'message': {
                'id': message.id,
                'name': message.name,
                'email': message.email,
                'phone': message.phone,
                'subject': message.subject,
                'message': message.message,
                'created_at': message.created_at.strftime('%B %d, %Y at %H:%M'),
                'is_seen': message.is_seen,
                'resolved': message.resolved,
            }
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def mark_message_seen(request, message_id):
    """Mark message as seen"""
    try:
        message = get_object_or_404(ContactMessage, id=message_id)
        message.is_seen = True
        message.save()
        return JsonResponse({
            'success': True,
            'message': 'Message marked as seen'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def mark_message_resolved(request, message_id):
    """Mark message as resolved"""
    try:
        message = get_object_or_404(ContactMessage, id=message_id)
        message.resolved = True
        message.save()
        return JsonResponse({
            'success': True,
            'message': 'Message marked as resolved'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_message(request, message_id):
    """Delete message"""
    try:
        message = get_object_or_404(ContactMessage, id=message_id)
        message.delete()
        return JsonResponse({
            'success': True,
            'message': 'Message deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Programs Management
@login_required
@user_passes_test(is_staff_member)
def programs_management(request):
    """Programs management"""
    programs = Program.objects.all().order_by('-created_at')
    active_programs_count = programs.filter(is_active=True).count()
    context = {'programs': programs, 'active_programs_count': active_programs_count}
    return render(request, 'staff_dashboard/programs.html', context)

@login_required
@user_passes_test(is_staff_member)
@require_POST
def create_program(request):
    """Create program via AJAX"""
    try:
        title = request.POST.get('title')
        description = request.POST.get('description')
        image = request.FILES.get('image')

        program = Program.objects.create(
            title=title,
            description=description,
            image=image,
            is_active=True
        )
        return JsonResponse({
            'success': True,
            'message': 'Program created successfully',
            'program_id': program.id
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def update_program(request, program_id):
    """Update program via AJAX"""
    try:
        program = get_object_or_404(Program, id=program_id)

        if request.content_type == 'application/json':
            # Handle JSON data (for toggle operations)
            import json
            data = json.loads(request.body)
            program.is_active = data.get('is_active', program.is_active)
        else:
            # Handle form data
            program.title = request.POST.get('title', program.title)
            program.description = request.POST.get('description', program.description)
            program.is_active = request.POST.get('is_active') == 'on'

            if 'image' in request.FILES:
                program.image = request.FILES['image']

        program.save()

        return JsonResponse({
            'success': True,
            'message': 'Program updated successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def toggle_program(request, program_id):
    """Toggle program active status via AJAX"""
    try:
        program = get_object_or_404(Program, id=program_id)
        data = json.loads(request.body)
        program.is_active = data.get('is_active', not program.is_active)
        program.save()
        return JsonResponse({
            'success': True,
            'message': f'Program {"activated" if program.is_active else "deactivated"} successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_program(request, program_id):
    """Delete program via AJAX"""
    try:
        program = get_object_or_404(Program, id=program_id)
        program.delete()
        return JsonResponse({
            'success': True,
            'message': 'Program deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Testimonials Management
@login_required
@user_passes_test(is_staff_member)
def testimonials_management(request):
    """Testimonials management"""
    testimonials = Testimonial.objects.all().order_by('-created_at')
    paginator = Paginator(testimonials, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'total_testimonials': testimonials.count(),
        'approved_count': testimonials.filter(status='approved').count(),
        'pending_count': testimonials.filter(status='pending').count(),
        'rejected_count': testimonials.filter(status='rejected').count(),
    }
    return render(request, 'staff_dashboard/testimonials.html', context)

@login_required
@user_passes_test(is_staff_member)
@require_POST
def approve_testimonial(request, testimonial_id):
    """Approve testimonial"""
    try:
        testimonial = get_object_or_404(Testimonial, id=testimonial_id)
        testimonial.status = 'approved'
        testimonial.save()
        return JsonResponse({
            'success': True,
            'message': 'Testimonial approved successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def reject_testimonial(request, testimonial_id):
    """Reject testimonial"""
    try:
        testimonial = get_object_or_404(Testimonial, id=testimonial_id)
        testimonial.status = 'rejected'
        testimonial.save()
        return JsonResponse({
            'success': True,
            'message': 'Testimonial rejected'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def delete_testimonial(request, testimonial_id):
    """Delete testimonial"""
    try:
        testimonial = get_object_or_404(Testimonial, id=testimonial_id)
        testimonial.delete()
        return JsonResponse({
            'success': True,
            'message': 'Testimonial deleted successfully'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
def export_donations(request):
    """Export donations data in various formats"""
    from django.http import HttpResponse
    import csv
    from openpyxl import Workbook
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib import colors
    from io import BytesIO

    # Get filtered donations
    donations = Donation.objects.all().order_by('-created_at')

    # Apply filters
    payment_method = request.GET.get('payment_method')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if payment_method:
        donations = donations.filter(payment_method=payment_method)

    if start_date:
        donations = donations.filter(created_at__date__gte=start_date)

    if end_date:
        donations = donations.filter(created_at__date__lte=end_date)

    format_type = request.GET.get('format', 'csv')

    if format_type == 'csv':
        # CSV Export
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="donations.csv"'

        writer = csv.writer(response)
        writer.writerow(['ID', 'Donor Name', 'Email', 'Amount', 'Currency', 'Payment Method', 'Status', 'Date', 'Transaction ID'])

        for donation in donations:
            writer.writerow([
                donation.id,
                donation.donor_name or 'Anonymous',
                donation.donor_email or '',
                donation.amount,
                donation.currency,
                donation.get_payment_method_display(),
                donation.get_status_display(),
                donation.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                donation.transaction_id or ''
            ])

        return response

    elif format_type == 'excel':
        # Excel Export
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="donations.xlsx"'

        wb = Workbook()
        ws = wb.active
        ws.title = "Donations"

        # Headers
        headers = ['ID', 'Donor Name', 'Email', 'Amount', 'Currency', 'Payment Method', 'Status', 'Date', 'Transaction ID']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Data
        for row_num, donation in enumerate(donations, 2):
            ws.cell(row=row_num, column=1, value=donation.id)
            ws.cell(row=row_num, column=2, value=donation.donor_name or 'Anonymous')
            ws.cell(row=row_num, column=3, value=donation.donor_email or '')
            ws.cell(row=row_num, column=4, value=float(donation.amount))
            ws.cell(row=row_num, column=5, value=donation.currency)
            ws.cell(row=row_num, column=6, value=donation.get_payment_method_display())
            ws.cell(row=row_num, column=7, value=donation.get_status_display())
            ws.cell(row=row_num, column=8, value=donation.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            ws.cell(row=row_num, column=9, value=donation.transaction_id or '')

        wb.save(response)
        return response

    elif format_type == 'pdf':
        # PDF Export
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="donations.pdf"'

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Title
        styles = getSampleStyleSheet()
        title = Paragraph("Donations Report", styles['Title'])
        elements.append(title)
        elements.append(Paragraph(" ", styles['Normal']))  # Spacer

        # Table data
        data = [['ID', 'Donor Name', 'Email', 'Amount', 'Currency', 'Payment Method', 'Status', 'Date']]

        for donation in donations:
            data.append([
                str(donation.id),
                donation.donor_name or 'Anonymous',
                donation.donor_email or '',
                f"${donation.amount:.2f}",
                donation.currency,
                donation.get_payment_method_display(),
                donation.get_status_display(),
                donation.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])

        # Create table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))

        elements.append(table)
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

    # Default to CSV if format not recognized
    return export_donations(request)

# Donations Management (View Only)
@login_required
@user_passes_test(is_staff_member)
def donations_management(request):
    """Donations management - view only"""
    donations = Donation.objects.all().order_by('-created_at')

    # Apply filters
    payment_method = request.GET.get('payment_method')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if payment_method:
        donations = donations.filter(payment_method=payment_method)

    if start_date:
        donations = donations.filter(created_at__date__gte=start_date)

    if end_date:
        donations = donations.filter(created_at__date__lte=end_date)

    paginator = Paginator(donations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Build query string for pagination links
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'total_donations': donations.count(),
        'total_amount': sum(d.amount for d in donations if d.amount),
        'query_string': query_string,
        'payment_method': payment_method,
        'start_date': start_date,
        'end_date': end_date,
        'payment_methods': Donation.PAYMENT_METHODS,
    }
    return render(request, 'staff_dashboard/donations.html', context)

@login_required
@user_passes_test(is_staff_member)
def view_donation(request, donation_id):
    """View donation details via AJAX"""
    print(f"view_donation called with donation_id: {donation_id}")
    print(f"User: {request.user}, is_authenticated: {request.user.is_authenticated}, is_staff_member: {request.user.is_staff_member()}")
    try:
        donation = get_object_or_404(Donation, id=donation_id)
        print(f"Found donation: {donation.id}")
        return JsonResponse({
            'success': True,
            'donation': {
                'id': donation.id,
                'donor_name': donation.donor_name,
                'donor_email': donation.donor_email,
                'donor_phone': donation.donor_phone,
                'amount': str(donation.amount),
                'currency': donation.currency,
                'status': donation.status,
                'status_display': donation.get_status_display(),
                'payment_method': donation.payment_method,
                'payment_method_display': donation.get_payment_method_display(),
                'transaction_id': donation.transaction_id,
                'message': donation.message,
                'created_at': donation.created_at.isoformat(),
                'receipt_number': getattr(donation, 'receipt_number', None),
                'mpesa_number': getattr(donation, 'mpesa_number', None),
                'card_last_four': getattr(donation, 'card_last_four', None),
                'paypal_transaction_id': getattr(donation, 'paypal_transaction_id', None),
                'payment_details': getattr(donation, 'payment_details', None),
            }
        })
    except Exception as e:
        print(f"Error in view_donation: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_staff_member)
@require_POST
def generate_receipt(request, donation_id):
    """Generate receipt for donation"""
    try:
        donation = get_object_or_404(Donation, id=donation_id)

        # Generate receipt number if not exists
        if not hasattr(donation, 'receipt_number') or not donation.receipt_number:
            import uuid
            donation.receipt_number = f"RCP-{donation.id}-{str(uuid.uuid4())[:8].upper()}"
            donation.save()

        # For now, return success - in a real implementation, you'd generate a PDF
        return JsonResponse({
            'success': True,
            'message': 'Receipt generated successfully',
            'receipt_number': donation.receipt_number,
            'receipt_url': f'/media/receipts/receipt_{donation.id}.pdf'  # Placeholder URL
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

# Backup Management
@login_required
@user_passes_test(is_staff_member)
def backup_management(request):
    """Backup management page"""
    context = {}
    return render(request, 'staff_dashboard/backup.html', context)
